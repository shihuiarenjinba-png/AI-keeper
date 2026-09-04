from __future__ import annotations

from datetime import datetime
from pathlib import Path
import csv
import shutil

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


ROOT = Path(__file__).resolve().parent
OUT = ROOT / "test_data"


def _write_csv(path: Path, rows: list[list[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["日付", "伝票番号", "数量", "金額"])
        writer.writerows(rows)


def _create_test_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "入力"

    ws.merge_cells("A1:H1")
    ws["A1"] = "CSV → Excel 安全転記 テスト用ブック"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = "年度"
    ws["B2"] = "2026年度"

    headers = ["日付", "伝票番号", "数量", "金額", "単価（数式）", "", "集計", "値"]
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    ws["A5"] = datetime(2026, 8, 28)
    ws["B5"] = "PRE001"
    ws["C5"] = 2
    ws["D5"] = 200
    ws["A6"] = datetime(2026, 8, 31)
    ws["B6"] = "PRE002"
    ws["C6"] = 3
    ws["D6"] = 450

    # A:D が転記対象。E列は数式保持確認用。
    for row in range(5, 51):
        ws[f"E{row}"] = f'=IFERROR(D{row}/C{row},"")'

    ws["G5"] = "既存＋追記件数"
    ws["G6"] = "合計数量"
    ws["G7"] = "合計金額"
    ws["G8"] = "平均単価"
    ws["H5"] = "=COUNTA(B5:B5000)"
    ws["H6"] = "=SUM(C5:C5000)"
    ws["H7"] = "=SUM(D5:D5000)"
    ws["H8"] = '=IFERROR(H7/H6,"")'

    for row in range(5, 5001):
        ws[f"A{row}"].number_format = "yyyy/mm/dd"
        ws[f"C{row}"].number_format = "0"
        ws[f"D{row}"].number_format = "#,##0"
        ws[f"E{row}"].number_format = "#,##0.00"

    widths = {"A": 14, "B": 16, "C": 10, "D": 14, "E": 16, "F": 3, "G": 16, "H": 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A5"
    wb.save(path)
    wb.close()


def main() -> None:
    if OUT.exists():
        shutil.rmtree(OUT)

    normal = OUT / "正常系"
    abnormal = OUT / "異常系"
    stress = OUT / "負荷テスト"
    normal.mkdir(parents=True)
    abnormal.mkdir(parents=True)
    stress.mkdir(parents=True)

    _create_test_workbook(normal / "2026年度_テスト.xlsx")

    _write_csv(
        normal / "2026_09_東京.csv",
        [
            ["2026/09/01", "T001", 1, 120],
            ["2026/09/02", "T002", 2, 260],
            ["2026/09/04", "T003", 3, 390],
        ],
    )
    _write_csv(
        normal / "2026_09_大阪.csv",
        [
            ["2026/09/01", "O001", 2, 220],
            ["2026/09/03", "O002", 4, 520],
            ["2026/09/05", "O003", 1, 140],
        ],
    )
    _write_csv(
        normal / "2026_09_名古屋.csv",
        [
            ["2026/09/02", "N001", 5, 650],
            ["2026/09/06", "N002", 2, 300],
        ],
    )

    _write_csv(abnormal / "2026_10_月違い.csv", [["2026/10/01", "X001", 1, 100]])
    _write_csv(abnormal / "2026_09_重複_A.csv", [["2026/09/07", "DUP001", 1, 100]])
    _write_csv(abnormal / "2026_09_重複_B.csv", [["2026/09/07", "DUP001", 2, 200]])
    _write_csv(
        abnormal / "2026_09_日付逆転.csv",
        [
            ["2026/09/10", "REV001", 1, 100],
            ["2026/09/09", "REV002", 1, 100],
        ],
    )

    rows: list[list[object]] = []
    index = 1
    for day in range(1, 31):
        count = 34 if day <= 10 else 33
        for _ in range(count):
            qty = (index % 9) + 1
            rows.append([f"2026/09/{day:02d}", f"S{index:04d}", qty, qty * 125])
            index += 1
    _write_csv(stress / "2026_09_ストレステスト_1000件.csv", rows)

    guide = """CSV → Excel 安全転記アプリ テストセット\n\n【正常系】\n1. 2026年度_テスト.xlsx をコピーしてから使います。\n2. 2026_09_東京.csv / 大阪.csv / 名古屋.csv を1行1パスで指定します。\n3. 2026年度_テスト.xlsx をExcel欄に指定します。\n4. 事前チェック → 更新を実行します。\n\n期待結果:\n- Excel既存データは5～6行目。\n- CSV 8件が7～14行目へ追記。\n- A:Dだけにデータが入り、E列とG:Hの数式は変更されない。\n- Excelと同じフォルダに _backup が作成される。\n\n【異常系】\n- 月違いを9月CSVと同時指定 → 停止。\n- 重複_A と 重複_B を同時指定 → 重複キーで停止。\n- 日付逆転CSV → 日付順エラーで停止。\n\n【負荷テスト】\n- 1000件CSVは正常系の負荷確認用です。\n- 必ず元のテストExcelをコピーし直してから実行してください。\n\n一度更新したExcelへ同じCSVを再投入して停止するのは正常な安全動作です。\n"""
    (OUT / "テスト手順.txt").write_text(guide, encoding="utf-8-sig")

    print(f"テストファイルを作成しました: {OUT}")


if __name__ == "__main__":
    main()
