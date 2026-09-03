from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from .config import AppConfig, ColumnSpec
from .csv_loader import CsvData, CsvRecord, normalize_key_part


@dataclass(frozen=True)
class PreflightReport:
    record_count: int
    min_date: str
    max_date: str
    fiscal_year: int
    existing_last_row: int
    existing_last_date: str | None
    target_start_row: int
    target_end_row: int
    formula_count: int
    warnings: tuple[str, ...] = ()


@dataclass(frozen=True)
class ProcessResult:
    workbook_bytes: bytes
    output_name: str
    record_count: int
    formula_count: int
    target_start_row: int
    target_end_row: int


def _is_formula(cell) -> bool:
    return cell.data_type == "f" or (
        isinstance(cell.value, str) and cell.value.startswith("=")
    )


def _formula_snapshot(workbook) -> dict[str, str]:
    snapshot: dict[str, str] = {}
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if _is_formula(cell):
                    snapshot[f"{ws.title}!{cell.coordinate}"] = str(cell.value)
    return snapshot


def _formula_diff(before: dict[str, str], after: dict[str, str]) -> list[str]:
    keys = sorted(set(before) | set(after))
    return [
        f"{key}: BEFORE={before.get(key)!r} / AFTER={after.get(key)!r}"
        for key in keys
        if before.get(key) != after.get(key)
    ]


def _excel_value_to_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y%m%d", "%Y年%m月%d日"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                pass
    return None


def _normalize_excel_value(value: Any, spec: ColumnSpec) -> str:
    if value in (None, ""):
        return ""
    if spec.data_type == "date":
        d = _excel_value_to_date(value)
        return d.isoformat() if d else f"<INVALID_DATE:{value}>"
    if spec.data_type == "number":
        try:
            return format(Decimal(str(value)).normalize(), "f")
        except Exception:
            return f"<INVALID_NUMBER:{value}>"
    return str(value).strip()


def _to_excel_value(value: Any, spec: ColumnSpec) -> Any:
    if value is None:
        return None
    if spec.data_type == "date":
        return datetime(value.year, value.month, value.day)
    if spec.data_type == "number":
        return float(value)
    return str(value)


class WorkbookEngine:
    def __init__(self, config: AppConfig):
        self.config = config

    def _load(self, workbook_bytes: bytes, workbook_name: str):
        suffix = Path(workbook_name).suffix.lower()
        if suffix not in {".xlsx", ".xlsm"}:
            raise ValueError("Excelは .xlsx または .xlsm を選択してください。")
        return load_workbook(
            BytesIO(workbook_bytes),
            data_only=False,
            keep_vba=suffix == ".xlsm",
            keep_links=True,
        )

    def _sheet(self, workbook):
        if self.config.sheet_name not in workbook.sheetnames:
            raise ValueError(f"Excelにシート '{self.config.sheet_name}' がありません。")
        return workbook[self.config.sheet_name]

    def _existing_state(self, ws) -> tuple[int, date | None]:
        col_idx = column_index_from_string(self.config.date_excel_column)
        last_row = self.config.data_start_row - 1
        last_date: date | None = None
        found_blank_after_data = False

        for row in range(self.config.data_start_row, self.config.data_end_row + 1):
            value = ws.cell(row=row, column=col_idx).value
            if value in (None, ""):
                if last_row >= self.config.data_start_row:
                    found_blank_after_data = True
                continue
            if found_blank_after_data:
                raise ValueError(
                    f"Excelの日付列 {self.config.date_excel_column} に途中の空欄があります。"
                    "追記位置を誤る可能性があるため停止します。"
                )
            parsed = _excel_value_to_date(value)
            if parsed is None:
                raise ValueError(
                    f"{ws.title}!{self.config.date_excel_column}{row} の日付を解釈できません。"
                )
            last_row = row
            last_date = parsed

        return last_row, last_date

    def _validate_fiscal_year(self, ws, fiscal_year: int) -> None:
        cell_ref = self.config.workbook_fiscal_year_cell
        if not cell_ref:
            return
        value = ws[cell_ref].value
        text = "" if value is None else str(value)
        match = re.search(r"(19|20)\d{2}", text)
        if not match:
            raise ValueError(f"年度確認セル {cell_ref} から年度を読み取れません: {text!r}")
        workbook_year = int(match.group(0))
        if workbook_year != fiscal_year:
            raise ValueError(
                f"CSVは {fiscal_year}年度ですが、Excelは {workbook_year}年度です。"
            )

    def _validate_target_empty(self, ws, start_row: int, end_row: int) -> None:
        for spec in self.config.columns:
            col_idx = column_index_from_string(spec.excel_column)
            for row in range(start_row, end_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if _is_formula(cell):
                    raise ValueError(
                        f"転記先 {ws.title}!{cell.coordinate} に数式があります。"
                        "数式セルには書き込みません。"
                    )
                if cell.value not in (None, ""):
                    raise ValueError(
                        f"転記先 {ws.title}!{cell.coordinate} に既存値があります。"
                        "上書きを防ぐため停止します。"
                    )

    def _existing_keys(self, ws, last_row: int) -> set[tuple[str, ...]]:
        if last_row < self.config.data_start_row:
            return set()
        specs = [self.config.spec_by_csv_header(h) for h in self.config.key_csv_headers]
        result: set[tuple[str, ...]] = set()
        for row in range(self.config.data_start_row, last_row + 1):
            parts: list[str] = []
            for spec in specs:
                col_idx = column_index_from_string(spec.excel_column)
                parts.append(_normalize_excel_value(ws.cell(row=row, column=col_idx).value, spec))
            result.add(tuple(parts))
        return result

    def _preflight_loaded(self, workbook, csv_data: CsvData, workbook_name: str):
        ws = self._sheet(workbook)
        self._validate_fiscal_year(ws, csv_data.fiscal_year)
        last_row, last_date = self._existing_state(ws)

        if (
            self.config.require_strictly_after_last_date
            and last_date is not None
            and csv_data.min_date <= last_date
        ):
            raise ValueError(
                f"Excel最終日 {last_date} に対し、CSV開始日 {csv_data.min_date} "
                "が後の日付ではありません。重複・上書き防止のため停止します。"
            )

        start_row = last_row + 1
        end_row = start_row + len(csv_data.records) - 1
        if end_row > self.config.data_end_row:
            raise ValueError(
                f"転記可能行が不足しています。必要最終行={end_row}, "
                f"設定上限={self.config.data_end_row}。自動で行挿入はしません。"
            )

        self._validate_target_empty(ws, start_row, end_row)
        existing_keys = self._existing_keys(ws, last_row)
        overlap = [r.key for r in csv_data.records if r.key in existing_keys]
        if overlap:
            raise ValueError(f"Excel既存データと重複するキーがあります。例: {overlap[:5]}")

        formulas = _formula_snapshot(workbook)
        warnings: list[str] = []
        if Path(workbook_name).suffix.lower() == ".xlsm":
            warnings.append(
                ".xlsm はVBAを保持する設定で保存しますが、ActiveX・特殊アドイン・外部接続などは"
                "実ファイルでの追加検証を推奨します。"
            )
        return ws, last_row, last_date, start_row, end_row, formulas, tuple(warnings)

    def preflight(self, csv_data: CsvData, workbook_bytes: bytes, workbook_name: str) -> PreflightReport:
        workbook = self._load(workbook_bytes, workbook_name)
        try:
            _, last_row, last_date, start_row, end_row, formulas, warnings = self._preflight_loaded(
                workbook, csv_data, workbook_name
            )
            return PreflightReport(
                record_count=len(csv_data.records),
                min_date=csv_data.min_date.isoformat(),
                max_date=csv_data.max_date.isoformat(),
                fiscal_year=csv_data.fiscal_year,
                existing_last_row=last_row,
                existing_last_date=last_date.isoformat() if last_date else None,
                target_start_row=start_row,
                target_end_row=end_row,
                formula_count=len(formulas),
                warnings=warnings,
            )
        finally:
            workbook.close()

    def _write_records(self, ws, records: tuple[CsvRecord, ...], start_row: int) -> None:
        for offset, record in enumerate(records):
            row = start_row + offset
            for spec in self.config.columns:
                col_idx = column_index_from_string(spec.excel_column)
                ws.cell(row=row, column=col_idx).value = _to_excel_value(
                    record.values[spec.csv_header], spec
                )

    def _validate_written(self, ws, records: tuple[CsvRecord, ...], start_row: int) -> list[str]:
        errors: list[str] = []
        for offset, record in enumerate(records):
            row = start_row + offset
            for spec in self.config.columns:
                col_idx = column_index_from_string(spec.excel_column)
                actual = ws.cell(row=row, column=col_idx).value
                actual_norm = _normalize_excel_value(actual, spec)
                expected_norm = normalize_key_part(record.values[spec.csv_header], spec)
                if actual_norm != expected_norm:
                    errors.append(
                        f"{ws.title}!{spec.excel_column}{row}: expected={expected_norm!r}, actual={actual_norm!r}"
                    )
                    if len(errors) >= 50:
                        return errors
        return errors

    def process(self, csv_data: CsvData, workbook_bytes: bytes, workbook_name: str) -> ProcessResult:
        workbook = self._load(workbook_bytes, workbook_name)
        try:
            ws, _, _, start_row, end_row, formulas_before, _ = self._preflight_loaded(
                workbook, csv_data, workbook_name
            )
            self._write_records(ws, csv_data.records, start_row)

            diff_after_write = _formula_diff(formulas_before, _formula_snapshot(workbook))
            if diff_after_write:
                raise RuntimeError(
                    "転記直後に数式変更を検出しました。\n" + "\n".join(diff_after_write[:20])
                )

            out = BytesIO()
            workbook.save(out)
            output_bytes = out.getvalue()
        finally:
            workbook.close()

        reopened = self._load(output_bytes, workbook_name)
        try:
            ws2 = self._sheet(reopened)
            diff_after_save = _formula_diff(formulas_before, _formula_snapshot(reopened))
            if diff_after_save:
                raise RuntimeError(
                    "保存後の再検証で数式変更を検出しました。\n" + "\n".join(diff_after_save[:20])
                )
            value_errors = self._validate_written(ws2, csv_data.records, start_row)
            if value_errors:
                raise RuntimeError(
                    "保存後の値照合で不一致を検出しました。\n" + "\n".join(value_errors[:20])
                )
        finally:
            reopened.close()

        source = Path(workbook_name)
        return ProcessResult(
            workbook_bytes=output_bytes,
            output_name=f"{source.stem}_転記済{source.suffix}",
            record_count=len(csv_data.records),
            formula_count=len(formulas_before),
            target_start_row=start_row,
            target_end_row=end_row,
        )
